#!/usr/bin/env python3
"""Display real outputs to verify system functionality."""

import sqlite3
import json
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None

def show_database_samples():
    """Show real usernames from database."""
    print("\n" + "="*80)
    print("DATABASE: Real Username Samples (showing regex normalization)")
    print("="*80 + "\n")
    
    conn = sqlite3.connect('clan_data.db')
    cursor = conn.cursor()
    
    # Get table names
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
    tables = [row[0] for row in cursor.fetchall()]
    print(f"Tables in database: {', '.join(tables)}\n")
    
    # Get columns in clan_members
    cursor.execute("PRAGMA table_info(clan_members)")
    cols = cursor.fetchall()
    print(f"clan_members columns: {[col[1] for col in cols]}\n")
    
    # Show real usernames
    cursor.execute("SELECT * FROM clan_members ORDER BY RANDOM() LIMIT 10")
    rows = cursor.fetchall()
    col_names = [description[0] for description in cursor.description]
    
    print("Sample Members (10 random):")
    print("-" * 100)
    for i, row in enumerate(rows, 1):
        print(f"{i}. ", end="")
        for name, val in zip(col_names[:5], row[:5]):  # Show first 5 columns
            if name == 'name':
                print(f"Name: '{val}'", end=" | ")
            else:
                print(f"{name}: {val}", end=" | ")
        print()
    
    conn.close()


def show_json_dashboard():
    """Show sample dashboard data."""
    print("\n" + "="*80)
    print("DASHBOARD: JSON Data Sample (clan_data.json)")
    print("="*80 + "\n")
    
    if Path('clan_data.json').exists():
        with open('clan_data.json', 'r') as f:
            data = json.load(f)
        
        print(f"File size: {Path('clan_data.json').stat().st_size / 1024:.1f} KB")
        print(f"Keys: {list(data.keys())}\n")
        
        # Show summary stats
        if 'total_members' in data:
            print(f"Total Members: {data['total_members']}")
        
        # Show activity heatmap structure
        if 'activity_heatmap' in data:
            heatmap = data['activity_heatmap']
            print(f"Activity Heatmap Days: {len(heatmap)} entries")
            if isinstance(heatmap, list):
                print(f"  Sample days: {heatmap[:2]}\n")
            elif isinstance(heatmap, dict):
                print(f"  Sample: {list(heatmap.items())[:2]}\n")
        
        # Show member data if available
        print("Top 5 Members by Boss Kills:")
        print("-" * 100)
        
        # Try different possible keys for member lists
        members = None
        member_key = None
        for key in ['allMembers', 'members', 'clan_members', 'stats']:
            if isinstance(data.get(key), list):
                members = data[key]
                member_key = key
                break
        
        if members:
            sorted_members = sorted(
                members, 
                key=lambda m: m.get('boss_kills') or m.get('kills') or 0, 
                reverse=True
            )[:5]
            
            for i, member in enumerate(sorted_members, 1):
                name = member.get('name') or member.get('username', 'N/A')
                kills = member.get('boss_kills') or member.get('kills', 0)
                msgs = member.get('message_count') or member.get('messages', 0)
                print(f"{i}. {name:<30} | Boss Kills: {kills:>6} | Messages: {msgs:>6}")
        else:
            print("(Member data structure differs - see keys above)")


def show_excel_report():
    """Show sample Excel report."""
    print("\n" + "="*80)
    print("EXCEL REPORT: Sample Data (clan_report_full.xlsx)")
    print("="*80 + "\n")
    
    excel_path = Path('clan_report_full.xlsx')
    if excel_path.exists():
        print(f"File size: {excel_path.stat().st_size / 1024:.2f} KB")
        
        if openpyxl:
            try:
                wb = openpyxl.load_workbook(excel_path, data_only=True)
                print(f"Sheets: {wb.sheetnames}")
                
                if 'Members' in wb.sheetnames:
                    ws = wb['Members']
                    print(f"\nMembers Sheet - First 6 rows:")
                    print("-" * 120)
                    
                    for i, row in enumerate(ws.iter_rows(max_row=6, values_only=True), 1):
                        if i == 1:  # Header
                            print(" | ".join(f"{str(v):<15}" for v in row[:8]))
                            print("-" * 120)
                        else:
                            print(" | ".join(f"{str(v):<15}" for v in row[:8]))
            except Exception as e:
                print(f"Could not read Excel: {e}")
        else:
            print("openpyxl not available - Excel file exists but cannot display content")


def show_app_log():
    """Show recent log entries."""
    print("\n" + "="*80)
    print("LOGS: Recent Pipeline Execution Trace")
    print("="*80 + "\n")
    
    log_path = Path('app.log')
    if log_path.exists():
        lines = log_path.read_text().strip().split('\n')
        print(f"Total log lines: {len(lines)}\n")
        print("Last 20 log entries (with trace IDs):")
        print("-" * 100)
        for line in lines[-20:]:
            print(line)


if __name__ == '__main__':
    try:
        show_database_samples()
        show_json_dashboard()
        show_excel_report()
        show_app_log()
        print("\n" + "="*80)
        print("END OF REAL OUTPUT SAMPLES")
        print("="*80 + "\n")
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
